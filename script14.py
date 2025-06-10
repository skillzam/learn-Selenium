# Selenium Python automation script :
# Reads test data from an Excel sheet
# Fills in the form at: https://demoqa.com/automation-practice-form
# Takes a screenshot for each test case
# Updates the Excel sheet with "Pass" or "Fail" status
# +------------------------------+
# |            Start             |
# +------------------------------+
#               |
#               v
# +------------------------------+
# | Initialize WebDriver & Excel |
# +------------------------------+
#               |
#               v
# +------------------------------+
# | Create 'screenshots' folder  |
# +------------------------------+
#               |
#               v
# +------------------------------+
# | Open URL: demoqa.com/form    |
# +------------------------------+
#               |
#               v
# +--------------------------------------+
# |   For each row in Excel (Test Case)  |
# +--------------------------------------+
#               |
#               v
# +--------------------------------------+
# |    Extract test data from the row    |
# +--------------------------------------+
#               |
#               v
# +--------------------------------+
# | Convert date format if needed  |
# +--------------------------------+
#               |
#               v
# +------------------------------+
# | Build absolute image path    |
# +------------------------------+
#             |
#             v
# +------------------------------+
# | Call fill_practice_form()    |
# +------------------------------+
#             |
#             v
# +------------------------------+
# | Was form filled successfully?|
# +----------+-------------------+
#            |Yes              |No
#            v                 v
# +---------------------+   +--------------------+
# | Set result = "Pass" |   | Set result = "Fail"|
# +---------------------+   +--------------------+
#            \                 /
#             v               v
#       +----------------------------+
#       | Take screenshot (by TC ID)|
#       +----------------------------+
#             |
#             v
# +----------------------------------+
# | Update result in Excel ("Pass/Fail") |
# +----------------------------------+
#             |
#             v
# +----------------------------+
# | Close modal & reload form |
# +----------------------------+
#             |
#             v
#      [Loop until all rows done]
#             |
#             v
# +----------------------------+
# | Save Excel workbook       |
# +----------------------------+
#             |
#             v
# +----------------------------+
# | Quit browser & End script |
# +----------------------------+

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime
import time
import os

chrome_options = Options()
chrome_options.add_argument("--start-maximized")
# chrome_options.add_argument("--headless")

# Paths
excel_path = 'test_data.xlsx'
screenshots_dir = 'screenshots'

# Create screenshots directory if not exists
os.makedirs(screenshots_dir, exist_ok=True)

# Initialize WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                          options=chrome_options)

driver.maximize_window()

def fill_practice_form(first_name, last_name, email, gender, mobile, dob,
                       subjects, hobbies, picture_path, address, state, city):
    try:
        # Fill form fields
        driver.find_element(By.ID, "firstName").send_keys(first_name)
        driver.find_element(By.ID, "lastName").send_keys(last_name)
        driver.find_element(By.ID, "userEmail").send_keys(email)

        # Gender selection
        gender_radio = driver.find_element(By.XPATH, f'//label[text()="{gender}"]')
        driver.execute_script("arguments[0].scrollIntoView();", gender_radio)
        gender_radio.click()

        # Mobile number
        driver.find_element(By.ID, "userNumber").send_keys(mobile)

        # Date of Birth - Clear and send formatted string
        dob_field = driver.find_element(By.ID, "dateOfBirthInput")
        dob_field.clear()
        dob_field.send_keys(dob)
        dob_field.send_keys(Keys.ENTER)

        # Subjects
        subjects_field = driver.find_element(By.ID, "subjectsInput")
        for subject in subjects.split(','):
            subjects_field.send_keys(subject.strip())
            subjects_field.send_keys(Keys.ENTER)

        # Hobbies
        for hobby in hobbies.split(','):
            hobby_checkbox = driver.find_element(By.XPATH, f'//label[text()="{hobby.strip()}"]')
            hobby_checkbox.click()

        # Upload Picture
        driver.find_element(By.ID, "uploadPicture").send_keys(picture_path)

        # Current Address
        driver.find_element(By.ID, "currentAddress").send_keys(address)

        # State and City
        driver.find_element(By.ID, "react-select-3-input").send_keys(state)
        driver.find_element(By.ID, "react-select-3-input").send_keys(Keys.ENTER)

        driver.find_element(By.ID, "react-select-4-input").send_keys(city)
        driver.find_element(By.ID, "react-select-4-input").send_keys(Keys.ENTER)

        # Submit Form
        driver.find_element(By.ID, "submit").click()

        # Wait for modal to appear
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, "example-modal-sizes-title-lg"))
        )
        return True

    except Exception as e:
        print(f"Error filling form: {e}")
        return False


# Load Workbook
wb = load_workbook(excel_path)
ws = wb.active

# Open the URL
driver.get("https://demoqa.com/automation-practice-form") 

# Loop through each row (skip header)
for row in ws.iter_rows(min_row=2, values_only=True):
    tc_id, fname, lname, email, gender, mobile, dob, subjects, hobbies, pic, address, state, city, _ = row

    print(f"\nRunning Test Case: {tc_id}")

    # Convert DOB from datetime to string if needed
    if isinstance(dob, datetime):
        dob = dob.strftime("%d %b %Y")  # Converts to "10 Jan 1990"

    # Construct absolute path for image
    pic_abs_path = os.path.abspath(pic) if pic else ""

    # Fill form
    result = fill_practice_form(fname, lname, email, gender, mobile, dob, subjects,
                                hobbies, pic_abs_path, address, state, city)

    # Take screenshot
    screenshot_path = os.path.join(screenshots_dir, f"{tc_id}.png")
    driver.save_screenshot(screenshot_path)

    # Update Excel with Pass/Fail
    for r in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row):
        if r[0].value == tc_id:
            ws[f"N{r[0].row}"] = "Pass" if result else "Fail"
            break

    # Close modal and clear form
    try:
        driver.find_element(By.ID, "closeLargeModal").click()
        driver.get("https://demoqa.com/automation-practice-form")   # Refresh page
    except:
        driver.get("https://demoqa.com/automation-practice-form") 

# Save updated Excel
wb.save(excel_path)
print("\nTest completed. Results saved in test_data.xlsx")

# Quit browser
driver.quit()
